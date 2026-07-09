#!/usr/bin/env python3
"""
build.py — regenerates index.html for the Used Car Campaigns dashboard.

HOW TO USE:
1. Replace/update "Used_cars_data.xlsx" in this same folder with your latest data.
2. Double-click this file (or run `python build.py` / `python3 build.py`
   from a terminal in this folder).
3. index.html will be regenerated. Open it locally to check, then
   commit + push to GitHub (see README.md for GitHub Desktop steps).

WHAT IT DOES:
- Reads Used_cars_data.xlsx (must have 'Reference', 'Google', 'Facebook' sheets)
- Excludes 'Demand Gen' and 'P.Max' campaign types
- Weeks before the last 7 days are shown as a daily-average bar (per week)
- The most recent ~week is shown day-by-day (see DAILY_CUTOFF_DAYS below)
- Injects the resulting JSON + date-range text into template.html
- Writes the final, self-contained index.html
"""

import sys

try:
    import openpyxl
except ImportError:
    print("=" * 60)
    print("ERROR: the 'openpyxl' package is not installed.")
    print()
    print("Open a terminal / command prompt in this folder and run:")
    print("    pip install openpyxl")
    print("(or 'pip3 install openpyxl' / 'python -m pip install openpyxl')")
    print("=" * 60)
    input("\nPress Enter to close...")
    sys.exit(1)

import json
import datetime
from collections import defaultdict
from pathlib import Path

# ---------------------------------------------------------------------------
# CONFIG — tweak these if your data / rules change
# ---------------------------------------------------------------------------
FOLDER = Path(__file__).parent
XLSX_PATH = FOLDER / "Used_cars_data.xlsx"
TEMPLATE_PATH = FOLDER / "template.html"
OUTPUT_PATH = FOLDER / "index.html"

EXCLUDE_TYPES = {"Demand Gen", "P.Max"}   # campaign types to drop entirely
DAILY_CUTOFF_DAYS = 7                     # last N days shown day-wise instead of weekly-avg


def get_city(campaign_name):
    c = campaign_name.lower()
    if "nashik" in c or "nasik" in c:
        return "Nashik"
    if "ahm" in c or "ahmedabad" in c:
        return "Ahmedabad"
    if "chd" in c or "chandigarh" in c:
        return "Chandigarh"
    return "Other"


def week_start(d):
    return d - datetime.timedelta(days=d.weekday())


def load_records(wb):
    ref = wb["Reference"]
    refmap = {}
    for row in ref.iter_rows(min_row=2, values_only=True):
        if row[0]:
            refmap[row[0]] = row[1]

    records = []
    for sheet_name in ["Google", "Facebook"]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            day, camp, cost, conv = row[0], row[1], row[2], row[3]
            if camp is None or day is None:
                continue
            ctype = refmap.get(camp, "Unknown")
            if ctype in EXCLUDE_TYPES:
                continue
            city = get_city(camp)
            if city == "Other":
                continue
            records.append({
                "date": day, "city": city, "type": ctype,
                "spend": float(cost or 0), "leads": float(conv or 0),
            })
    return records


def build_dataset(records):
    max_date = max(r["date"] for r in records)
    daily_cutoff = max_date - datetime.timedelta(days=DAILY_CUTOFF_DAYS - 1)
    daily_cutoff = week_start(daily_cutoff)  # snap to a Monday for clean week boundaries

    cities = sorted(set(r["city"] for r in records))
    types = sorted(set(r["type"] for r in records))

    # ---- per (city, type) weekly + daily series ----
    aggw = defaultdict(lambda: defaultdict(float))
    daysw = defaultdict(set)
    for r in records:
        if r["date"] >= daily_cutoff:
            continue
        key = (r["city"], r["type"], week_start(r["date"]))
        aggw[key]["spend"] += r["spend"]
        aggw[key]["leads"] += r["leads"]
        daysw[key].add(r["date"].date())
    weeks = sorted(set(k[2] for k in aggw))

    aggd = defaultdict(lambda: defaultdict(float))
    for r in records:
        if r["date"] < daily_cutoff:
            continue
        key = (r["city"], r["type"], r["date"])
        aggd[key]["spend"] += r["spend"]
        aggd[key]["leads"] += r["leads"]
    days = sorted(set(k[2] for k in aggd))

    data = {}
    for city in cities:
        data[city] = {}
        for t in types:
            points = []
            for w in weeks:
                key = (city, t, w)
                if key in aggw:
                    n_days = len(daysw[key]) or 1
                    spend_avg = round(aggw[key]["spend"] / n_days, 2)
                    leads_avg = round(aggw[key]["leads"] / n_days, 2)
                    cpl = round(spend_avg / leads_avg, 2) if leads_avg > 0 else None
                    points.append({"label": w.strftime("%d %b"), "granularity": "week",
                                   "spend": spend_avg, "leads": leads_avg, "cpl": cpl, "days": n_days})
            for d in days:
                key = (city, t, d)
                if key in aggd:
                    spend = round(aggd[key]["spend"], 2)
                    leads = round(aggd[key]["leads"], 2)
                    cpl = round(spend / leads, 2) if leads > 0 else None
                    points.append({"label": d.strftime("%d %b"), "granularity": "day",
                                   "spend": spend, "leads": leads, "cpl": cpl, "days": 1})
            if points:
                data[city][t] = points

    # ---- Total series (all types combined), computed independently from raw records ----
    aggw_t = defaultdict(lambda: defaultdict(float))
    daysw_t = defaultdict(set)
    for r in records:
        if r["date"] >= daily_cutoff:
            continue
        key = (r["city"], week_start(r["date"]))
        aggw_t[key]["spend"] += r["spend"]
        aggw_t[key]["leads"] += r["leads"]
        daysw_t[key].add(r["date"].date())
    weeks_t = sorted(set(k[1] for k in aggw_t))

    aggd_t = defaultdict(lambda: defaultdict(float))
    for r in records:
        if r["date"] < daily_cutoff:
            continue
        key = (r["city"], r["date"])
        aggd_t[key]["spend"] += r["spend"]
        aggd_t[key]["leads"] += r["leads"]
    days_t = sorted(set(k[1] for k in aggd_t))

    total_data = {}
    for city in cities:
        series = []
        for w in weeks_t:
            key = (city, w)
            if key in aggw_t:
                n_days = len(daysw_t[key]) or 1
                spend_avg = round(aggw_t[key]["spend"] / n_days, 2)
                leads_avg = round(aggw_t[key]["leads"] / n_days, 2)
                cpl = round(spend_avg / leads_avg, 2) if leads_avg > 0 else None
                series.append({"label": w.strftime("%d %b"), "granularity": "week",
                               "spend": spend_avg, "leads": leads_avg, "cpl": cpl})
        for d in days_t:
            key = (city, d)
            if key in aggd_t:
                spend = round(aggd_t[key]["spend"], 2)
                leads = round(aggd_t[key]["leads"], 2)
                cpl = round(spend / leads, 2) if leads > 0 else None
                series.append({"label": d.strftime("%d %b"), "granularity": "day",
                               "spend": spend, "leads": leads, "cpl": cpl})
        total_data[city] = series

    # ---- Monthly totals (for the June-vs-May-style KPI strip) ----
    month_names = {1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
                   7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December"}
    aggm = defaultdict(lambda: defaultdict(float))
    for r in records:
        key = (r["city"], r["date"].month)
        aggm[key]["spend"] += r["spend"]
        aggm[key]["leads"] += r["leads"]
    monthly = {}
    for city in cities:
        months = sorted(set(k[1] for k in aggm if k[0] == city))
        pts = []
        for m in months:
            key = (city, m)
            spend = round(aggm[key]["spend"], 2)
            leads = round(aggm[key]["leads"], 2)
            cpl = round(spend / leads, 2) if leads > 0 else None
            pts.append({"month": month_names[m], "spend": spend, "leads": leads, "cpl": cpl})
        monthly[city] = pts

    min_date = min(r["date"] for r in records)
    range_text = f"{min_date.strftime('%d %b')} – {max_date.strftime('%d %b %Y')}"
    daily_range_text = f"{daily_cutoff.strftime('%d %b')} – {max_date.strftime('%d %b')}"

    return data, total_data, monthly, range_text, daily_range_text


def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"ERROR: {XLSX_PATH.name} not found in this folder. "
                          f"Put your data file next to build.py and try again.")
    if not TEMPLATE_PATH.exists():
        raise SystemExit(f"ERROR: {TEMPLATE_PATH.name} not found. "
                          f"This file should not be deleted/renamed.")

    print(f"Reading {XLSX_PATH.name} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    records = load_records(wb)
    print(f"  -> {len(records)} rows loaded across "
          f"{len(set(r['city'] for r in records))} cities, "
          f"{len(set(r['type'] for r in records))} campaign types")

    data, total_data, monthly, range_text, daily_range_text = build_dataset(records)

    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    output = (template
              .replace("__DATA_JSON__", json.dumps(data, separators=(",", ":")))
              .replace("__MONTHLY_JSON__", json.dumps(monthly, separators=(",", ":")))
              .replace("__TOTAL_JSON__", json.dumps(total_data, separators=(",", ":")))
              .replace("__RANGE_TEXT__", range_text)
              .replace("__DAILY_RANGE_TEXT_UPPER__", daily_range_text.upper())
              .replace("__DAILY_RANGE_TEXT__", daily_range_text))

    OUTPUT_PATH.write_text(output, encoding="utf-8")
    print(f"  -> wrote {OUTPUT_PATH.name} ({len(output):,} bytes)")
    print(f"  -> date range: {range_text}  |  daily section: {daily_range_text}")
    print("Done. Open index.html to check it, then commit + push in GitHub Desktop.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nBUILD FAILED: {e}")
    finally:
        input("\nPress Enter to close...")
